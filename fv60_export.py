import os
import re
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def normalize_store_key(value) -> str:
    s = str(value).strip()
    if not s or s.lower() == "nan":
        return ""

    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""

    if len(digits) == 4:
        return digits.zfill(5)

    if len(digits) == 5:
        return digits

    return digits


def load_cost_center_mapping(filepath: str) -> dict:
    if not os.path.exists(filepath):
        return {}

    try:
        df = pd.read_excel(filepath, dtype=str)
    except Exception:
        return {}

    df = df.fillna("")
    cols = {str(c).strip().lower(): c for c in df.columns}

    store_col = None
    legacy_store_col = None
    cost_col = None
    profit_col = None

    for key, original in cols.items():
        if "store" in key or "branch" in key or "สาขา" in key:
            store_col = original
            break

    for key, original in cols.items():
        if "legacy profit center code" in key:
            legacy_store_col = original
            break

    for key, original in cols.items():
        if "cost center" in key or key == "cost center" or "cost" in key:
            cost_col = original
            break

    for key, original in cols.items():
        if "profit center" in key and "legacy" not in key:
            profit_col = original
            break
    if not profit_col:
        for key, original in cols.items():
            if "profit" in key and "legacy" not in key:
                profit_col = original
                break

    print("MAPPING DEBUG => headers =", list(df.columns))
    print("MAPPING DEBUG => store_col =", store_col)
    print("MAPPING DEBUG => legacy_store_col =", legacy_store_col)
    print("MAPPING DEBUG => cost_col =", cost_col)
    print("MAPPING DEBUG => profit_col =", profit_col)

    mapping = {}

    for _, row in df.iterrows():
        cost_center_val = str(row.get(cost_col, "")).strip() if cost_col else ""
        profit_center_val = str(row.get(profit_col, "")).strip() if profit_col else ""

        record = {
            "cost_center": cost_center_val,
            "profit_center": profit_center_val,
        }

        candidate_keys = []

        if store_col:
            candidate_keys.append(str(row.get(store_col, "")).strip())

        if legacy_store_col:
            candidate_keys.append(str(row.get(legacy_store_col, "")).strip())

        if cost_center_val:
            m = re.search(r"(\d{4,5})$", cost_center_val)
            if m:
                candidate_keys.append(m.group(1))

        normalized_keys = []
        for raw in candidate_keys:
            k = normalize_store_key(raw)
            if k:
                normalized_keys.append(k)

        normalized_keys = list(dict.fromkeys(normalized_keys))

        for k in normalized_keys:
            mapping[k] = record
            mapping[k.lstrip("0")] = record

    print("MAPPING DEBUG => total keys =", len(mapping))
    print("MAPPING DEBUG => sample 01266 =", mapping.get("01266"))
    print("MAPPING DEBUG => sample 1266 =", mapping.get("1266"))
    print("MAPPING DEBUG => sample 83224 =", mapping.get("83224"))
    print("MAPPING DEBUG => sample 83226 =", mapping.get("83226"))

    return mapping


def autosize_worksheet(ws):
    for column_cells in ws.columns:
        length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            try:
                value_len = len(str(cell.value)) if cell.value is not None else 0
                length = max(length, value_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(length + 2, 40)


def style_excel(output_path: str):
    wb = load_workbook(output_path)
    ws = wb.active

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    autosize_worksheet(ws)
    wb.save(output_path)


def build_dynamic_headers():
    return [
        "Company Code",      # A
        "Vendor",            # B
        "B",                 # C
        "Invoice Date",      # D
        "Reference",         # E
        "Posting Date",      # F
        "Amount",            # G
        "Currency",          # H
        "W/H",               # I
        "Text",              # J
        "Exchange Rate",     # K
        "Tax",               # L
        "Tax Code",          # M
        "Business places",   # N
        "Section",           # O
        "Text",              # P
        "GST Partner",       # Q
        "POS",               # R
        "General Ledger",    # S
        "D.C",               # T
        "Amount",            # U
        "Assignment",        # V
        "Text",              # W
        "Cost Center",       # X
        "Order",             # Y
        "Profit Center",     # Z
        "Payment Term",      # AA
        "Header Text",       # AB
        "Ref Key1",          # AC
        "Ref Key 2",         # AD
        "Ref Key 3",         # AE
        "HSN.SAC",           # AF
        "Branch Number",     # AG
        "Base Amount",       # AH
        "Payment Method",    # AI
        "Partner Bank",      # AJ
        "House Bank",        # AK
        "Account ID",        # AL
        "Baseline Date",     # AM
    ]


def row_to_dynamic_values(item: dict) -> list:
    return [
        item.get("company_code", "7590"),          # A
        item.get("vendor", ""),                    # B
        item.get("b", ""),                         # C
        item.get("invoice_date", ""),              # D
        item.get("reference", ""),                 # E
        item.get("posting_date", ""),              # F
        item.get("amount", 0.0),                   # G
        item.get("currency", "THB"),               # H
        item.get("w/h", ""),                       # I
        item.get("text", ""),                      # J
        item.get("exchange_rate", ""),             # K
        item.get("tax", "X"),                      # L
        item.get("tax_code", "I3"),                # M
        item.get("business_place", "7590"),        # N
        item.get("section", ""),                   # O
        item.get("text", ""),                      # P
        item.get("gst_partner", ""),               # Q
        item.get("pos", ""),                       # R
        item.get("general_ledger", "41511101"),    # S
        item.get("d.c", "S"),                      # T
        item.get("amount", 0.0),                   # U
        item.get("assignment", "Vendor Invoice"),  # V
        item.get("text", ""),                      # W
        item.get("cost_center", ""),               # X
        item.get("order", ""),                     # Y
        item.get("profit_center", ""),             # Z
        item.get("payment_term", "0001"),          # AA
        item.get("header_text", ""),               # AB
        item.get("ref_key1", ""),                  # AC
        item.get("ref_key2", ""),                  # AD
        item.get("ref_key3", ""),                  # AE
        item.get("hsn.sac", ""),                   # AF
        item.get("company_code", "7590"),          # AG
        item.get("base_amount", ""),               # AH
        item.get("payment_method", ""),            # AI
        item.get("partner_bank", ""),              # AJ
        item.get("house_bank", ""),                # AK
        item.get("account_id", ""),                # AL
        item.get("baseline_date", ""),             # AM
    ]


def export_dynamic_fv60_excel(rows: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "FV60_Output"

    headers = build_dynamic_headers()
    ws.append(headers)

    for item in rows:
        ws.append(row_to_dynamic_values(item))

    wb.save(output_path)
    style_excel(output_path)